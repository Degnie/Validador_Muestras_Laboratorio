import pandas as pd
from fastapi.testclient import TestClient

from app.api.muestras import _build_reporte_excel
from app.core.config import Settings
from app.main import create_app


def _write_dataset(dir_path):
    pd.DataFrame(
        {
            "id_muestra": ["M-001", "M-001"],
            "tipo_analisis": ["Agua Potable", "Agua Potable"],
            "prueba_requerida": ["pH", "Metales_Pesados"],
        }
    ).to_excel(dir_path / "Checklist_Maestro.xlsx", index=False)
    with pd.ExcelWriter(dir_path / "Datos.xlsx") as writer:
        pd.DataFrame(
            {
                "ID": ["M-0O1"],
                "Resultado": ["OK"],
                "Valor": ["7.2"],
                "Tecnico que realizo": ["Tec. Pérez"],
                "Fecha": ["2026-07-11"],
            }
        ).to_excel(writer, sheet_name="pH", index=False)


def test_get_muestras_returns_status_with_fuzzy_matched_id(tmp_path):
    _write_dataset(tmp_path)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras")

    assert response.status_code == 200
    body = response.json()
    muestras = {m["id_muestra"]: m for m in body["muestras"]}
    assert muestras["M-001"]["estado"] == "Faltante"
    assert muestras["M-001"]["tipo_analisis"] == "Agua Potable"
    assert muestras["M-001"]["pruebas_faltantes"] == ["Metales_Pesados"]
    assert muestras["M-001"]["pruebas"] == [
        {"nombre_prueba": "pH", "resultado": "OK", "valor": "7.2", "tecnico": "Tec. Pérez", "fecha": "2026-07-11"}
    ]


def test_get_muestras_reports_partial_success_when_a_checklist_row_is_invalid(tmp_path):
    # Una fila rota en el checklist ya no aborta el lote entero (422): se descarta esa fila,
    # se sigue procesando el resto, y el detalle queda en errores_validacion.
    _write_dataset(tmp_path)
    pd.DataFrame(
        {
            "id_muestra": ["M-001", "M-002"],
            "tipo_analisis": ["Agua Potable", "Agua Potable"],
            "prueba_requerida": ["pH", None],
        }
    ).to_excel(tmp_path / "Checklist_Maestro.xlsx", index=False)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras")

    assert response.status_code == 200
    body = response.json()
    assert "M-002" not in {m["id_muestra"] for m in body["muestras"]}
    assert len(body["errores_validacion"]) == 1
    assert "Fila 1" in body["errores_validacion"][0]
    # Mensaje para un técnico, no jerga de pydantic ("Input should be a valid string"): el
    # nombre de columna legible + un código en vez del texto crudo de la excepción. Una celda
    # vacía en un Excel llega como NaN (no como clave ausente), así que pydantic la reporta
    # como tipo inválido (VAL-002), no como campo faltante (VAL-001).
    assert "prueba requerida" in body["errores_validacion"][0]
    assert "VAL-002" in body["errores_validacion"][0]
    assert "Input" not in body["errores_validacion"][0]


def test_buscar_returns_only_matching_codes(tmp_path):
    _write_dataset(tmp_path)
    pd.DataFrame(
        {
            "id_muestra": ["M-001", "M-001", "M-777", "M-777"],
            "tipo_analisis": ["Agua Potable"] * 4,
            "prueba_requerida": ["pH", "Metales_Pesados", "pH", "Metales_Pesados"],
        }
    ).to_excel(tmp_path / "Checklist_Maestro.xlsx", index=False)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras/buscar", params={"q": "M-001"})

    assert response.status_code == 200
    ids = {m["id_muestra"] for m in response.json()["muestras"]}
    assert ids == {"M-001"}


def test_buscar_with_blank_query_returns_no_results(tmp_path):
    _write_dataset(tmp_path)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras/buscar", params={"q": ""})

    assert response.status_code == 200
    assert response.json()["muestras"] == []


def test_exportar_returns_a_readable_xlsx(tmp_path):
    _write_dataset(tmp_path)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras/exportar")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    import io

    hojas = pd.read_excel(io.BytesIO(response.content), sheet_name=None)
    assert set(hojas.keys()) == {"Resumen", "Detalle_pruebas"}

    resumen = hojas["Resumen"]
    assert list(resumen.columns) == ["ID", "Estado", "Tipo_analisis", "Pruebas_faltantes", "Pruebas_adicionales"]
    assert "M-001" in resumen["ID"].values
    fila = resumen[resumen["ID"] == "M-001"].iloc[0]
    assert fila["Pruebas_faltantes"] == "Metales_Pesados"  # texto plano, no una lista de Python

    detalle = hojas["Detalle_pruebas"]
    assert list(detalle.columns) == ["ID", "Prueba", "Resultado", "Valor", "Tecnico", "Fecha"]
    assert "pH" in detalle.loc[detalle["ID"] == "M-001", "Prueba"].values


def test_exportar_formats_headers_and_widens_columns(tmp_path):
    # El técnico no debe tener que agrandar las columnas a mano ni adivinar dónde empieza el
    # encabezado -- ver el pedido explícito en el CHANGELOG de esta iteración.
    import io

    from openpyxl import load_workbook

    _write_dataset(tmp_path)
    app = create_app(Settings(data_dir=tmp_path))
    client = TestClient(app)

    response = client.get("/api/muestras/exportar")

    wb = load_workbook(io.BytesIO(response.content))
    for nombre_hoja in ("Resumen", "Detalle_pruebas"):
        ws = wb[nombre_hoja]
        encabezado = ws[1][0]
        assert encabezado.font.bold is True
        assert encabezado.fill.start_color.rgb == "000E6E63"
        for columna in ws.column_dimensions.values():
            assert columna.width is not None and columna.width >= 10
        assert ws.freeze_panes == "A2"


def test_exportar_alertas_pendientes_returns_a_formatted_xlsx():
    import io

    from openpyxl import load_workbook

    app = create_app(Settings())
    client = TestClient(app)

    response = client.post(
        "/api/notificaciones/exportar",
        json={
            "alertas": [
                {"id_muestra": "M-011", "prueba": "Dureza_Total", "creada": "2026-07-19T10:00:00.000+00:00"},
            ]
        },
    )

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    df = pd.read_excel(io.BytesIO(response.content), sheet_name="Alertas_pendientes")
    assert list(df.columns) == ["ID", "Prueba_pendiente", "Alerta_creada"]
    assert df.iloc[0]["ID"] == "M-011"
    assert df.iloc[0]["Alerta_creada"] == "19/07/2026 10:00"

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Alertas_pendientes"]
    assert ws[1][0].font.bold is True


def test_exportar_alertas_pendientes_with_no_alerts_returns_only_the_header():
    app = create_app(Settings())
    client = TestClient(app)

    response = client.post("/api/notificaciones/exportar", json={"alertas": []})

    assert response.status_code == 200
    import io

    df = pd.read_excel(io.BytesIO(response.content), sheet_name="Alertas_pendientes")
    assert df.empty
    assert list(df.columns) == ["ID", "Prueba_pendiente", "Alerta_creada"]


def test_build_reporte_excel_relabels_pruebas_fantasma_and_flattens_lists():
    # El técnico que abre el Excel no debe ver el nombre interno del contrato ("Pruebas
    # Fantasma") ni literales de Python ([]/{}) en ninguna celda.
    estados = pd.DataFrame(
        [
            {
                "id_muestra": "M-009",
                "estado": "Pruebas Fantasma",
                "tipo_analisis": "Agua Potable",
                "pruebas_faltantes": ["Turbidez"],
                "pruebas_fantasma": ["Nitratos"],
                "pruebas": [
                    {"nombre_prueba": "pH", "resultado": "OK", "valor": "7.2", "tecnico": "Tec. Ríos", "fecha": "2026-07-11"},
                    {
                        "nombre_prueba": "Nitratos",
                        "resultado": "OK",
                        "valor": "5 mg/L",
                        "tecnico": "Tec. Ríos",
                        "fecha": "2026-07-11",
                    },
                ],
            }
        ]
    )

    resumen, detalle = _build_reporte_excel(estados)

    fila = resumen.iloc[0]
    assert fila["Estado"] == "Pruebas Adicionales"
    assert fila["Pruebas_faltantes"] == "Turbidez"
    assert fila["Pruebas_adicionales"] == "Nitratos"
    for valor in [*resumen.to_dict(orient="records")[0].values(), *detalle.to_dict(orient="records")[0].values()]:
        assert "[" not in str(valor) and "{" not in str(valor)

    assert list(detalle["Prueba"]) == ["Nitratos (adicional)", "Turbidez", "pH"]


def test_registrar_notificacion_appends_a_row_to_the_audit_csv(tmp_path):
    settings = Settings(data_dir=tmp_path)
    app = create_app(settings)
    client = TestClient(app)

    response = client.post("/api/notificaciones", json={"id_muestra": "M-001", "prueba": "Microbiologia"})

    assert response.status_code == 204
    lineas = settings.notificaciones_csv_path.read_text(encoding="utf-8").splitlines()
    assert lineas[0] == "id_muestra,prueba,fecha_deteccion"
    assert lineas[1].startswith("M-001,Microbiologia,")
