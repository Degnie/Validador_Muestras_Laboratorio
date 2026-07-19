"""Formato compartido para todo Excel que se le entrega a un técnico (no a otro programa):
encabezado que se distinga del contenido a simple vista y columnas anchas según su propio
contenido, para que no haga falta agrandarlas a mano cada vez que se abre el archivo."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# Mismo verde-azulado que --color-primary en frontend/src/main.css: el Excel se ve parte de
# la misma identidad visual que el dashboard.
_HEADER_FILL = PatternFill(start_color="0E6E63", end_color="0E6E63", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_BORDE = Border(*(Side(style="thin", color="B7C2BC") for _ in range(4)))

_ANCHO_MIN = 10
_ANCHO_MAX = 60
_PADDING = 4


def formatear_hoja(ws: Worksheet) -> None:
    for celda in ws[1]:
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = _HEADER_ALIGN
        celda.border = _BORDE

    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.border = _BORDE

    for columna in ws.columns:
        largo_max = max((len(str(celda.value)) for celda in columna if celda.value is not None), default=_ANCHO_MIN)
        letra = columna[0].column_letter
        ws.column_dimensions[letra].width = min(max(largo_max + _PADDING, _ANCHO_MIN), _ANCHO_MAX)

    ws.freeze_panes = "A2"
