import csv
import logging
import unicodedata
import zipfile
from pathlib import Path
from typing import TypeVar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from app.core.column_aliases import COLUMN_ALIASES
from app.core.error_codes import (
    ARCH_CONTENIDO,
    ARCH_EXTENSION,
    ARCH_SOSPECHOSO,
    ARCH_TAMANO,
    CAMPO_LEGIBLE,
    VAL_DATO_INVALIDO,
    VAL_FALTA_DATO,
    ErrorAplicacion,
)

logger = logging.getLogger(__name__)

MAX_EXCEL_SIZE_MB = 20
# Un .xlsx es un zip; openpyxl en read_only sí streamea las filas de la hoja, pero igual
# carga sharedStrings.xml completo en memoria para poder resolver los IDs de string de cada
# celda. Ahí es donde pega un Zip Bomb real: un part que pesa KB comprimido pero se
# descomprime a GB. Se detecta leyendo los tamaños del directorio central del zip (sin
# descomprimir nada) antes de dejar que openpyxl abra el archivo.
MAX_UNCOMPRESSED_MB = 200
MAX_COMPRESSION_RATIO = 100
ModelT = TypeVar("ModelT", bound=BaseModel)

# .xlsx is a zip archive; every real one starts with this signature. Checking it (instead of
# trusting the .xlsx extension alone) is what actually stops a renamed/malicious file from
# reaching the xlsx parser.
XLSX_MAGIC_BYTES = b"PK\x03\x04"


def _normalize_key(name: str) -> str:
    stripped = str(name).strip().lower()
    no_accents = unicodedata.normalize("NFKD", stripped).encode("ascii", "ignore").decode()
    return no_accents.replace(" ", "").replace("_", "")


def _canonical_column(name: str) -> str:
    key = _normalize_key(name)
    if key in COLUMN_ALIASES:
        return COLUMN_ALIASES[key]
    # ponytail: no alias on file -> best-effort snake_case, add alias entries as new variants appear
    return str(name).strip().lower().replace(" ", "_")


def assert_safe_excel_file(path: Path, max_size_mb: float = MAX_EXCEL_SIZE_MB) -> None:
    """Rejects files that aren't a plain .xlsx or that are suspiciously large before we parse
    them. El mensaje que lleva cada excepción ya es el texto para el técnico (ver
    app/core/error_codes.py); el detalle técnico real (tamaños, ratios) queda en el log del
    servidor, no en la respuesta al cliente."""
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        logger.warning("%s: extensión %s en %s", ARCH_EXTENSION.codigo, path.suffix, path.name)
        raise ErrorAplicacion(ARCH_EXTENSION)
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > max_size_mb:
        logger.warning("%s: %.2f MB (máximo %s MB) en %s", ARCH_TAMANO.codigo, size_mb, max_size_mb, path.name)
        raise ErrorAplicacion(ARCH_TAMANO)
    with open(path, "rb") as f:
        header = f.read(len(XLSX_MAGIC_BYTES))
    if header != XLSX_MAGIC_BYTES:
        logger.warning("%s: magic bytes no coinciden en %s", ARCH_CONTENIDO.codigo, path.name)
        raise ErrorAplicacion(ARCH_CONTENIDO)

    try:
        with zipfile.ZipFile(path) as archivo_zip:
            total_uncompressed = sum(info.file_size for info in archivo_zip.infolist())
            total_compressed = sum(info.compress_size for info in archivo_zip.infolist()) or 1
    except zipfile.BadZipFile:
        logger.warning("%s: BadZipFile en %s", ARCH_CONTENIDO.codigo, path.name)
        raise ErrorAplicacion(ARCH_CONTENIDO) from None

    if total_uncompressed > MAX_UNCOMPRESSED_MB * 1024 * 1024:
        logger.warning(
            "%s: %.2f MB descomprimidos (máximo %s MB) en %s -- posible Zip Bomb",
            ARCH_SOSPECHOSO.codigo,
            total_uncompressed / (1024 * 1024),
            MAX_UNCOMPRESSED_MB,
            path.name,
        )
        raise ErrorAplicacion(ARCH_SOSPECHOSO)
    if total_uncompressed / total_compressed > MAX_COMPRESSION_RATIO:
        logger.warning(
            "%s: ratio de compresión %.0fx (máximo %sx) en %s -- posible Zip Bomb",
            ARCH_SOSPECHOSO.codigo,
            total_uncompressed / total_compressed,
            MAX_COMPRESSION_RATIO,
            path.name,
        )
        raise ErrorAplicacion(ARCH_SOSPECHOSO)


def _read_sheet_normalized(sheet: Worksheet, batch_size: int) -> pd.DataFrame:
    """Streams one worksheet in batches, instead of loading it whole into memory."""
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    columns = [_canonical_column(c) for c in header] if header else []

    batches: list[pd.DataFrame] = []
    batch: list[tuple] = []
    for row in rows:
        batch.append(row)
        if len(batch) >= batch_size:
            batches.append(pd.DataFrame(batch, columns=columns))
            batch = []
    if batch:
        batches.append(pd.DataFrame(batch, columns=columns))

    df = pd.concat(batches, ignore_index=True) if batches else pd.DataFrame(columns=columns)
    for col in df.select_dtypes(include=["object", "str"]).columns:
        df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)
    return df


def read_excel_normalized(path: Path, batch_size: int = 500) -> pd.DataFrame:
    """Streams the active sheet via openpyxl's read_only mode (no full-workbook object graph
    in memory) and assembles the DataFrame in batches, instead of pandas.read_excel loading
    everything at once."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _read_sheet_normalized(workbook.active, batch_size)
    finally:
        workbook.close()


def read_excel_multisheet_normalized(path: Path, batch_size: int = 500) -> dict[str, pd.DataFrame]:
    """Same streaming/batching approach as read_excel_normalized, but for every sheet of a
    multi-tab workbook (each sheet = one 'prueba'/area). Returns {nombre_hoja: DataFrame}."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {name: _read_sheet_normalized(workbook[name], batch_size) for name in workbook.sheetnames}
    finally:
        workbook.close()


def validate_rows(df: pd.DataFrame, schema: type[ModelT]) -> tuple[pd.DataFrame, list[str]]:
    """Runs every row through a Pydantic schema (type coercion + required fields). A bad row
    is skipped and reported, not fatal for the rest of the batch (partial success) -- a typo
    in one row of a 500-row Excel shouldn't block every other valid row.

    El mensaje de cada fila descartada usa CAMPO_LEGIBLE + un código (VAL-001/VAL-002, ver
    app/core/error_codes.py) en vez del texto crudo de pydantic ("Input should be a valid
    string") -- quien lee esto es un técnico de laboratorio, no otro programador."""
    validated = []
    errores = []
    for i, row in enumerate(df.to_dict(orient="records")):
        try:
            validated.append(schema(**row).model_dump())
        except ValidationError as exc:
            primer_error = exc.errors()[0]
            campo = ".".join(str(p) for p in primer_error["loc"])
            campo_legible = CAMPO_LEGIBLE.get(campo, campo)
            info = VAL_FALTA_DATO if primer_error["type"] == "missing" else VAL_DATO_INVALIDO
            mensaje = info.mensaje_usuario.format(campo=campo_legible)
            errores.append(f"Fila {i}: {mensaje} (código {info.codigo})")
    result = pd.DataFrame(validated, columns=list(schema.model_fields.keys()))
    return result, errores


def check_file_freshness(paths: dict[str, Path], max_lag_days: float = 1) -> list[str]:
    """Returns area names whose file is more than max_lag_days older than the most recent one."""
    mtimes = {area: Path(p).stat().st_mtime for area, p in paths.items()}
    newest = max(mtimes.values())
    max_lag_seconds = max_lag_days * 86400
    return [area for area, mtime in mtimes.items() if newest - mtime > max_lag_seconds]


NOTIFICACIONES_CSV_HEADER = ["id_muestra", "prueba", "fecha_deteccion"]


def append_notificacion_csv(csv_path: Path, id_muestra: str, prueba: str, fecha_deteccion: str) -> None:
    """Appends one audit row, writing the header first if the file doesn't exist yet.
    Sync by design (plain stdlib csv, no aiofiles dependency); the caller offloads this to a
    thread (asyncio.to_thread) so it doesn't block the event loop."""
    csv_path = Path(csv_path)
    is_new_file = not csv_path.exists()
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if is_new_file:
            writer.writerow(NOTIFICACIONES_CSV_HEADER)
        writer.writerow([id_muestra, prueba, fecha_deteccion])
